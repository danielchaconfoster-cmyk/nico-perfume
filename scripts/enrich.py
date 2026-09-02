import openpyxl
import json
import re
import os
import random

FAMILY_KEYWORDS = {
    'Oriental / Ámbar': ['amber', 'ámbar', 'oriental', 'oud', 'vanilla', 'vainilla', 'tonka', 'incienso', 'incense', 'resina', 'khamrah', 'untold', 'paragon', 'rouge', 'golden', 'gold', 'velvet', 'sugar', 'elixir', 'rich', 'baccarat', 'musk', 'almizcle'],
    'Amaderada': ['wood', 'cedro', 'cedar', 'sándalo', 'sandalwood', 'vetiver', 'guaiac', 'ebony', 'oud', 'patchouli', 'pachuli', 'king', 'terred', 'terre', 'woods', 'forest', 'noir', 'black', 'cypress'],
    'Cítrica / Fresca': ['citrus', 'cítrico', 'cítrica', 'bergamot', 'bergamota', 'lemon', 'limón', 'mandarin', 'mandarina', 'grapefruit', 'pomelo', 'aqua', 'acua', 'water', 'marine', 'mar', 'fresh', 'fresco', 'fresca', 'summer', 'bleu', 'blue', 'ck one', 'light blue', 'ocean'],
    'Floral': ['rose', 'rosa', 'jasmine', 'jazmín', 'flor', 'floral', 'iris', 'orquídea', 'orchid', 'tuberosa', 'neroli', 'magnolia', 'peonía', 'peony', 'violet', 'violeta', 'panthera', 'bloom', 'blossom', 'bouquet', 'chloe', 'good girl', 'libre', 'vie est belle', 'miss'],
    'Gourmand / Dulce': ['caramel', 'caramelo', 'chocolate', 'cafe', 'coffee', 'praline', 'praliné', 'miel', 'honey', 'marshmallow', 'gourmand', 'dulce', 'candy', 'sweet', 'pistachio', 'pistacho', 'yara', 'lattafa', 'vanilla'],
    'Cuero / Especiada': ['leather', 'cuero', 'spicy', 'especias', 'pimienta', 'pepper', 'canela', 'cinnamon', 'cardamom', 'cardamomo', 'tobacco', 'tabaco', 'wanted', 'stronger', 'sauvage', 'man', 'intense', 'spice'],
    'Aromática / Fougère': ['lavender', 'lavanda', 'salvia', 'sage', 'menta', 'mint', 'romero', 'rosemary', 'albahaca', 'basil', 'fougere', 'fougère', 'sport', 'club', 'classic', 'homme', 'pour homme', 'legend', 'eros', 'dylan']
}

NOTE_SETS = {
    'Oriental / Ámbar': {
        'top': ['Bergamota de Calabria', 'Azafrán', 'Cardamomo', 'Manzana crujiente', 'Canela de Ceilán', 'Naranja amarga'],
        'heart': ['Ámbar gris', 'Jazmín de Egipto', 'Rosa de Damasco', 'Resina de Benjuí', 'Incienso', 'Nuez moscada'],
        'base': ['Vainilla de Madagascar', 'Madera de Oud', 'Haba Tonka', 'Almizcle blanco', 'Cedro', 'Pachulí']
    },
    'Amaderada': {
        'top': ['Pomelo amargo', 'Pimienta rosa', 'Cardamomo', 'Bergamota', 'Elemi', 'Ciprés'],
        'heart': ['Cedro del Atlas', 'Vetiver de Haití', 'Madera de Guayaco', 'Pachulí', 'Iris', 'Geranio'],
        'base': ['Sándalo de Mysore', 'Musgo de roble', 'Ámbar negro', 'Cuero suave', 'Benjuí', 'Almizcle amaderado']
    },
    'Cítrica / Fresca': {
        'top': ['Bergamota italiana', 'Limón de Amalfi', 'Mandarina jugosa', 'Menta fresca', 'Notas marinas', 'Pomelo'],
        'heart': ['Neroli', 'Jazmín acuático', 'Jengibre', 'Cardamomo verde', 'Romero', 'Calone'],
        'base': ['Almizcle blanco', 'Cedro blanco', 'Vetiver suave', 'Ámbar gris cristalino', 'Maderas flotantes']
    },
    'Floral': {
        'top': ['Pera Williams', 'Mandarina dulce', 'Pimienta rosa', 'Grosellas negras', 'Flor de azahar'],
        'heart': ['Rosa de Mayo', 'Jazmín Sambac', 'Tuberosa de la India', 'Iris de Florencia', 'Peonía rosa'],
        'base': ['Vainilla bourbon', 'Almizcle sedoso', 'Sándalo cremoso', 'Haba tonka', 'Pachulí refinado']
    },
    'Gourmand / Dulce': {
        'top': ['Almendra amarga', 'Frutos rojos', 'Caramelo salado', 'Pera madura', 'Licor de cereza'],
        'heart': ['Vainilla de Tahití', 'Praliné', 'Chocolate negro', 'Café tostado', 'Jazmín dulce'],
        'base': ['Haba Tonka tostada', 'Ámbar dulce', 'Sándalo', 'Almizcle de malvavisco', 'Miel dorada']
    },
    'Cuero / Especiada': {
        'top': ['Pimienta negra', 'Cardamomo', 'Azafrán persa', 'Canela', 'Mandarina especiada'],
        'heart': ['Cuero toscano', 'Tabaco cubano', 'Incienso de Omán', 'Clavo de olor', 'Cacao'],
        'base': ['Madera de abedul', 'Pachulí oscuro', 'Ámbar resinoso', 'Vainilla ahumada', 'Vetiver oscuro']
    },
    'Aromática / Fougère': {
        'top': ['Lavanda francesa', 'Menta piperita', 'Bergamota', 'Manzana verde', 'Cardamomo'],
        'heart': ['Salvia esclarea', 'Geranio bourbon', 'Pimienta de Sichuan', 'Hojas de violeta'],
        'base': ['Haba tonka', 'Musgo de roble', 'Cedro de Virginia', 'Vetiver', 'Ambroxan']
    }
}

KNOWN_PERFUMES = {
    '9 PM': {
        'family': 'Oriental / Ámbar',
        'top': ['Manzana salvaje', 'Canela', 'Lavanda silvestre', 'Bergamota'],
        'heart': ['Flor de azahar', 'Lirio de los valles', 'Vainilla cálida'],
        'base': ['Haba tonka', 'Vainilla', 'Ámbar', 'Pachulí'],
        'vibe': 'Seductor, Nocturno, Dulce, Fiesta',
        'longevity': '9/10 (10-12 hrs)',
        'sillage': 'Muy Alta (Pesada)'
    },
    'Club de Nuit': {
        'family': 'Amaderada',
        'top': ['Limón', 'Grosellas negras', 'Manzana', 'Bergamota', 'Piña'],
        'heart': ['Rosa', 'Jazmín', 'Abedul'],
        'base': ['Vainilla', 'Ámbar gris', 'Almizcle', 'Pachulí'],
        'vibe': 'Poderoso, Elegante, Proyectante, Cumplidos',
        'longevity': '9.5/10 (12+ hrs)',
        'sillage': 'Enorme'
    },
    'King': {
        'family': 'Oriental / Ámbar',
        'top': ['Naranja dulce', 'Bergamota', 'Limón de Sicilia'],
        'heart': ['Notas frutales exóticas', 'Fruta de la pasión'],
        'base': ['Vainilla de Madagascar', 'Ámbar blanco', 'Almizcle'],
        'vibe': 'Avasallador, Juvenil, Dulce Frutal, Atractivo',
        'longevity': '10/10 (14+ hrs)',
        'sillage': 'Bomba Proyectante'
    },
    'Amber Oud': {
        'family': 'Oriental / Ámbar',
        'top': ['Bergamota', 'Notas verdes', 'Melón', 'Piña'],
        'heart': ['Gourmand dulce', 'Ámbar', 'Cedro'],
        'base': ['Vainilla', 'Almizcle', 'Resinas', 'Maderas nobles'],
        'vibe': 'Ultra Lujo, Rico, Exótico, Sofisticado',
        'longevity': '9.5/10 (12+ hrs)',
        'sillage': 'Muy Alta'
    },
    'Wanted': {
        'family': 'Cuero / Especiada',
        'top': ['Limón', 'Jengibre', 'Lavanda', 'Menta'],
        'heart': ['Cardamomo guatemalteco', 'Enebro', 'Manzana', 'Geranio'],
        'base': ['Vetiver de Haití', 'Haba tonka', 'Amberwood'],
        'vibe': 'Magnético, Masculino, Vibrante, Nocturno',
        'longevity': '8.5/10 (8-10 hrs)',
        'sillage': 'Alta'
    },
    'CK One': {
        'family': 'Cítrica / Fresca',
        'top': ['Limón', 'Notas verdes', 'Bergamota', 'Piña', 'Mandarina'],
        'heart': ['Lirio de los valles', 'Jazmín', 'Violeta', 'Rosa'],
        'base': ['Almizcle', 'Cedro', 'Sándalo', 'Musgo de roble', 'Ámbar'],
        'vibe': 'Limpio, Icónico, Diario, Refrescante',
        'longevity': '6.5/10 (5-6 hrs)',
        'sillage': 'Moderada / Íntima'
    }
}

LUXURY_BOTTLE_IMAGES = [
    'https://images.unsplash.com/photo-1594035910387-fea47794261f?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1523293182086-7651a899d37f?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1547887537-6158d64c35b3?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1592945403244-b3fbafd7f539?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1588405748880-12d1d2a59f75?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1615397349754-cfa2066a298e?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1587017539504-67cfbddac569?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1595425970377-c9703cf48b6d?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1583445013765-46c20c4a6772?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1563178406-4cdc2923acbc?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1508746829417-e6f548d8d6ed?w=800&auto=format&fit=crop&q=80',
    'https://images.unsplash.com/photo-1541643600914-78b084683601?w=800&auto=format&fit=crop&q=80'
]

def determine_family(name, brand):
    combined = f'{name} {brand}'.lower()
    for fam, kws in FAMILY_KEYWORDS.items():
        if any(kw in combined for kw in kws):
            return fam
    if 'mujer' in combined or 'woman' in combined or 'femme' in combined:
        return 'Floral'
    elif 'hombre' in combined or 'man' in combined or 'homme' in combined:
        return 'Amaderada'
    return 'Oriental / Ámbar'

def extract_volume(name):
    match = re.search(r'(\d+)\s*(?:ml|ML)', name)
    if match:
        return int(match.group(1))
    return 100

def extract_concentration(name):
    name_upper = name.upper()
    if 'EXTRAIT' in name_upper:
        return 'Extrait de Parfum'
    elif 'EDP' in name_upper or 'EAU DE PARFUM' in name_upper:
        return 'Eau de Parfum'
    elif 'EDT' in name_upper or 'EAU DE TOILETTE' in name_upper:
        return 'Eau de Toilette'
    elif 'EDC' in name_upper or 'COLOGNE' in name_upper:
        return 'Eau de Cologne'
    elif 'PARFUM' in name_upper or 'ELIXIR' in name_upper:
        return 'Parfum'
    return 'Eau de Parfum'

def clean_gender(raw_gender, name):
    raw = (raw_gender or '').lower()
    name_l = name.lower()
    if 'unisex' in raw or 'unisex' in name_l:
        return 'Unisex'
    elif 'mujer' in raw or 'woman' in raw or 'dama' in raw or 'femme' in name_l:
        return 'Mujer'
    elif 'hombre' in raw or 'man' in raw or 'homme' in name_l:
        return 'Hombre'
    return 'Unisex'

def clean_title(name, brand):
    cleaned = re.sub(r'^Perfume\s+', '', name, flags=re.IGNORECASE).strip()
    return cleaned

def generate_notes(name, family, brand):
    for k, v in KNOWN_PERFUMES.items():
        if k.lower() in name.lower():
            return v['top'], v['heart'], v['base'], v['vibe'], v['longevity'], v['sillage']
    
    seed = sum(ord(c) for c in name)
    rng = random.Random(seed)
    notes = NOTE_SETS[family]
    top = rng.sample(notes['top'], min(3, len(notes['top'])))
    heart = rng.sample(notes['heart'], min(3, len(notes['heart'])))
    base = rng.sample(notes['base'], min(3, len(notes['base'])))
    
    vibes = [
        'Elegante, Sofisticado, Firma Personal',
        'Seductor, Nocturno, Enigmático',
        'Fresco, Enérgico, Versátil para Diario',
        'Cálido, Acogedor, Sensual',
        'Poderoso, Proyectante, Alta Distinción',
        'Magnético, Vibrante, Moderno'
    ]
    vibe = vibes[seed % len(vibes)]
    longevities = ['8.5/10 (8-10 hrs)', '9/10 (10-12 hrs)', '9.5/10 (12+ hrs)', '7.5/10 (6-8 hrs)']
    sillages = ['Alta', 'Moderada a Alta', 'Enorme', 'Proyectante']
    return top, heart, base, vibe, longevities[seed % len(longevities)], sillages[seed % len(sillages)]

wb = openpyxl.load_workbook('lista 20 de agosto2026 - 3.xlsx', data_only=True)
sheet = wb['LISTA DE PRECIOS']

perfumes = []
seen_skus = set()

for r in range(14, sheet.max_row + 1):
    c1 = sheet.cell(row=r, column=1).value
    c2 = sheet.cell(row=r, column=2).value
    c3 = sheet.cell(row=r, column=3).value
    c4 = sheet.cell(row=r, column=4).value
    c5 = sheet.cell(row=r, column=5).value
    c6 = sheet.cell(row=r, column=6).value
    c7 = sheet.cell(row=r, column=7).value
    
    if not c2 or not isinstance(c7, (int, float)) or c7 <= 0:
        continue
        
    brand = str(c1).strip() if c1 else 'Prestige Perfumes'
    raw_name = str(c2).strip()
    sku = str(c3).strip() if c3 else f'PERF-{r}'
    
    if sku in seen_skus:
        sku = f'{sku}-{r}'
    seen_skus.add(sku)
    
    ean = str(c4).strip() if c4 else ''
    gender = clean_gender(c5, raw_name)
    formato = str(c6).strip() if c6 else 'REGULAR'
    wholesale = int(c7)
    
    retail_price = round(wholesale * 1.48 / 100) * 100
    normal_price = round(retail_price * 1.30 / 100) * 100
    
    volume = extract_volume(raw_name)
    concentration = extract_concentration(raw_name)
    family = determine_family(raw_name, brand)
    title = clean_title(raw_name, brand)
    
    top_notes, heart_notes, base_notes, vibe, longevity, sillage = generate_notes(raw_name, family, brand)
    img_idx = sum(ord(c) for c in raw_name) % len(LUXURY_BOTTLE_IMAGES)
    image_url = LUXURY_BOTTLE_IMAGES[img_idx]
    
    rating = round(4.5 + ((sum(ord(c) for c in raw_name) % 5) / 10), 1)
    review_count = 12 + (sum(ord(c) for c in raw_name) % 180)
    
    occasions = []
    if 'Nocturno' in vibe or 'Seductor' in vibe:
        occasions.extend(['Noche', 'Citas / Romance', 'Eventos Especiales'])
    if 'Fresco' in vibe or 'Diario' in vibe:
        occasions.extend(['Uso Diario', 'Oficina', 'Primavera / Verano'])
    if 'Poderoso' in vibe or 'Elegante' in vibe:
        occasions.extend(['Reuniones', 'Otoño / Invierno', 'Firma'])
    if not occasions:
        occasions = ['Versátil', 'Todo el año', 'Uso Diario']
        
    is_best_seller = (r % 7 == 0) or (brand in ['Afnan', 'Al Haramain', 'Bharara', 'Armaf', 'Azzaro', 'Calvin Klein', 'Cartier', 'Carolina Herrera', 'Lattafa'])
    is_new = (r % 11 == 0)
    
    item = {
        'id': f'perf-{r}',
        'sku': sku,
        'ean': ean,
        'name': title,
        'fullName': raw_name,
        'brand': brand,
        'gender': gender,
        'format': formato,
        'volume': volume,
        'concentration': concentration,
        'family': family,
        'price': retail_price,
        'originalPrice': normal_price,
        'wholesalePrice': wholesale,
        'stock': 15 + (r % 45),
        'topNotes': top_notes,
        'heartNotes': heart_notes,
        'baseNotes': base_notes,
        'allNotes': list(set(top_notes + heart_notes + base_notes)),
        'vibe': vibe,
        'longevity': longevity,
        'sillage': sillage,
        'occasions': occasions,
        'rating': rating,
        'reviews': review_count,
        'image': image_url,
        'isBestSeller': is_best_seller,
        'isNew': is_new,
        'description': f'{title} de {brand} es una creación olfativa excepcional de la familia {family}. Abre con notas de {", ".join(top_notes)}, evoluciona hacia un corazón de {", ".join(heart_notes)}, y asienta en una base de {", ".join(base_notes)}.'
    }
    perfumes.append(item)

os.makedirs('src/data', exist_ok=True)
with open('src/data/perfumes.json', 'w', encoding='utf-8') as f:
    json.dump(perfumes, f, ensure_ascii=False, indent=2)

brands = sorted(list(set(p['brand'] for p in perfumes)))
families = sorted(list(set(p['family'] for p in perfumes)))
genders = sorted(list(set(p['gender'] for p in perfumes)))

meta = {
    'total': len(perfumes),
    'brands': brands,
    'families': families,
    'genders': genders,
    'priceRange': {
        'min': min(p['price'] for p in perfumes),
        'max': max(p['price'] for p in perfumes)
    }
}

with open('src/data/meta.json', 'w', encoding='utf-8') as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

print(f'Done: {len(perfumes)} perfumes processed.')
